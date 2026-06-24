"""
Excel 导出模块
"""
import os
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# 列定义：(标题, 字段名, 列宽)
COLUMNS = [
    ("序号", "index", 8),
    ("用户昵称", "user_name", 20),
    ("评论内容", "comment_text", 60),
    ("点赞数", "like_count", 12),
    ("回复数", "reply_count", 12),
    ("发布时间", "publish_time", 22),
    ("评论ID", "comment_id", 20),
    ("用户ID", "user_id", 18),
]


def export_to_excel(
    comments: List[Dict],
    output_path: str,
    video_url: str = "",
) -> str:
    """
    将评论列表导出为 Excel 文件

    Args:
        comments: 评论列表
        output_path: 输出文件路径
        video_url: 视频链接（写入表头）

    Returns:
        实际保存的文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "评论数据"

    # 样式定义
    header_font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Microsoft YaHei", size=11)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 写入视频信息行
    if video_url:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        info_cell = ws.cell(row=1, column=1, value=f"视频链接: {video_url}")
        info_cell.font = Font(name="Microsoft YaHei", size=11, italic=True, color="666666")
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
        start_row = 3
    else:
        start_row = 1

    # 写入统计行
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(COLUMNS))
    stat_cell = ws.cell(
        row=start_row, column=1,
        value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    评论总数: {len(comments)}"
    )
    stat_cell.font = Font(name="Microsoft YaHei", size=11, italic=True, color="666666")
    stat_cell.alignment = Alignment(horizontal="left", vertical="center")
    header_row = start_row + 2

    # 写入表头
    for col_idx, (title, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # 写入数据
    for row_idx, comment in enumerate(comments, 1):
        data_row = header_row + row_idx
        for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
            if field == "index":
                value = row_idx
            else:
                value = comment.get(field, "")

            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border

            # 数值列居中
            if field in ("index", "like_count", "reply_count"):
                cell.alignment = center_alignment
                # 尝试转为数字
                if field != "index":
                    try:
                        cell.value = int(value)
                    except (ValueError, TypeError):
                        pass
            else:
                cell.alignment = cell_alignment

    # 自动筛选
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(comments)}"

    # 保存
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def get_default_output_path() -> str:
    """获取默认输出路径"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"抖音评论_{timestamp}.xlsx"
    return os.path.join(os.path.expanduser("~"), "Desktop", filename)
